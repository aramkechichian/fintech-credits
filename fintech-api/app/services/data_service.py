"""
Data export service for credit requests
Handles Excel export functionality
"""
import logging
from typing import List, Dict, Any, Optional
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.credit_request import CreditRequestInDB, CreditRequestStatus
from app.services.credit_request_service import search_credit_requests

logger = logging.getLogger(__name__)

# Available fields for export (based on CreditRequestResponse model)
# Only include fields that are actually used
AVAILABLE_FIELDS = {
    "id": "ID",
    "country": "País",
    "currency_code": "Código de Moneda",
    "full_name": "Nombre Completo",
    "email": "Email",
    "identity_document": "Documento de Identidad",
    "requested_amount": "Monto Solicitado",
    "monthly_income": "Ingreso Mensual",
    "request_date": "Fecha de Solicitud",
    "status": "Estado",
    "created_at": "Fecha de Creación",
    "updated_at": "Fecha de Actualización",
}

# Fields that are nested in bank_information (only if needed)
BANK_FIELDS = []


async def export_credit_requests_to_excel(
    countries: Optional[List[str]] = None,
    status: Optional[str] = None,
    request_date_from: Optional[datetime] = None,
    request_date_to: Optional[datetime] = None,
    selected_fields: Optional[List[str]] = None
) -> BytesIO:
    """
    Export credit requests to Excel based on filters and selected fields
    
    Args:
        countries: Optional list of countries to filter
        status: Optional status filter
        request_date_from: Optional start date for request_date filter
        request_date_to: Optional end date for request_date filter
        selected_fields: List of field names to include in export
        
    Returns:
        BytesIO: Excel file as bytes
    """
    try:
        # If no fields selected, use all available fields
        if not selected_fields:
            selected_fields = list(AVAILABLE_FIELDS.keys())
        
        # Validate selected fields - ONLY use fields that are in selected_fields
        valid_fields = [f for f in selected_fields if f in AVAILABLE_FIELDS]
        if not valid_fields:
            raise ValueError("No valid fields selected for export")
        
        # Log exactly what fields will be exported
        logger.info(f"Exporting with ONLY these selected fields: {valid_fields}")
        
        # Get all matching requests (no pagination for export)
        requests, total_count = await search_credit_requests(
            countries=countries,
            status=status,
            request_date_from=request_date_from,
            request_date_to=request_date_to,
            skip=0,
            limit=10000  # Large limit for export
        )
        
        # Check if there are any requests to export
        if total_count == 0:
            raise ValueError("No data found matching the selected filters")
        
        logger.info(f"Exporting {total_count} credit requests with fields: {valid_fields}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes de Crédito"
        
        # Create header row - ONLY for selected fields
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # ONLY create headers for the selected fields
        headers = [AVAILABLE_FIELDS[field] for field in valid_fields]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows - ONLY for selected fields
        for row_idx, request in enumerate(requests, start=2):
            for col_idx, field in enumerate(valid_fields, start=1):
                value = _get_field_value(request, field)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(valid_fields) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        logger.info(f"Excel file created successfully with {total_count} rows")
        return excel_file
        
    except Exception as e:
        logger.error(f"Error exporting credit requests to Excel: {str(e)}", exc_info=True)
        raise


def _get_field_value(request: CreditRequestInDB, field: str) -> Any:
    """
    Extract field value from credit request
    
    Handles nested fields like bank_information fields
    """
    if field == "id":
        return str(request.id)
    elif field == "country":
        return request.country.value if hasattr(request.country, 'value') else str(request.country)
    elif field == "currency_code":
        return request.currency_code.value if hasattr(request.currency_code, 'value') else str(request.currency_code)
    elif field == "full_name":
        return request.full_name
    elif field == "email":
        return request.email
    elif field == "identity_document":
        return request.identity_document
    elif field == "requested_amount":
        return request.requested_amount
    elif field == "monthly_income":
        return request.monthly_income
    elif field == "request_date":
        return request.request_date.isoformat() if request.request_date else ""
    elif field == "status":
        return request.status.value if hasattr(request.status, 'value') else str(request.status)
    elif field == "created_at":
        return request.created_at.isoformat() if request.created_at else ""
    elif field == "updated_at":
        return request.updated_at.isoformat() if request.updated_at else ""
    else:
        # Field not found - return empty string
        logger.warning(f"Field '{field}' not found in request, returning empty string")
        return ""


def get_available_fields() -> Dict[str, str]:
    """
    Get list of available fields for export
    
    Returns:
        Dict mapping field names to display labels
    """
    return AVAILABLE_FIELDS.copy()
