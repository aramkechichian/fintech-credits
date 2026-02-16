"""
Log export service for exporting logs to Excel
"""
import logging
from typing import List, Optional
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.log_data import LogDataInDB
from app.repositories.log_data_repository import log_data_repository
from app.utils.endpoint_mapper import get_module_name_for_endpoint

logger = logging.getLogger(__name__)

# Available fields for export
# Note: Module names will be in English for Excel export
# Frontend will handle translations when displaying
AVAILABLE_FIELDS = {
    "id": "ID",
    "endpoint": "Endpoint",
    "module": "Module",
    "method": "Method",
    "user_id": "User ID",
    "response_status": "Response Status",
    "is_success": "Is Success",
    "error_message": "Error Message",
    "created_at": "Created At",
}

# Module name mapping for Excel export (English)
MODULE_NAMES_EN = {
    "creditRequests": "Credit Request",
    "countryRules": "Country Rules",
    "authentication": "Authentication",
    "bankProvider": "Bank Provider",
    "audits": "Audits",
    "logs": "Logs",
}


async def export_logs_to_excel(
    method: Optional[str] = None,
    endpoint: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    selected_fields: Optional[List[str]] = None
) -> BytesIO:
    """
    Export logs to Excel based on filters and selected fields
    
    Args:
        method: Optional HTTP method filter
        endpoint: Optional endpoint filter
        date_from: Optional start date for created_at filter
        date_to: Optional end date for created_at filter
        selected_fields: List of field names to include in export
        
    Returns:
        BytesIO: Excel file as bytes
    """
    try:
        # If no fields selected, use all available fields
        if not selected_fields:
            selected_fields = list(AVAILABLE_FIELDS.keys())
        
        # Validate selected fields
        valid_fields = [f for f in selected_fields if f in AVAILABLE_FIELDS]
        if not valid_fields:
            raise ValueError("No valid fields selected for export")
        
        # Get all matching logs (no pagination for export)
        logs, total_count = await log_data_repository.search(
            method=method,
            endpoint=endpoint,
            date_from=date_from,
            date_to=date_to,
            skip=0,
            limit=10000  # Large limit for export
        )
        
        # Check if there are any logs to export
        if total_count == 0:
            raise ValueError("No data found matching the selected filters")
        
        logger.info(f"Exporting {total_count} logs with fields: {valid_fields}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"
        
        # Create header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = [AVAILABLE_FIELDS[field] for field in valid_fields]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_idx, log in enumerate(logs, start=2):
            for col_idx, field in enumerate(valid_fields, start=1):
                value = _get_field_value(log, field)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(valid_fields) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        logger.info(f"Excel file created successfully with {total_count} rows")
        return excel_file
        
    except Exception as e:
        logger.error(f"Error exporting logs to Excel: {str(e)}", exc_info=True)
        raise


def _get_field_value(log: LogDataInDB, field: str) -> any:
    """
    Extract field value from log entry
    """
    if field == "id":
        return str(log.id)
    elif field == "endpoint":
        return log.endpoint
    elif field == "module":
        module_key = get_module_name_for_endpoint(log.endpoint)
        if module_key:
            # Return English name for Excel export
            return MODULE_NAMES_EN.get(module_key, module_key)
        return log.endpoint
    elif field == "method":
        return log.method
    elif field == "user_id":
        return str(log.user_id) if log.user_id else ""
    elif field == "response_status":
        return log.response_status if log.response_status else ""
    elif field == "is_success":
        return "Yes" if log.is_success else "No"
    elif field == "error_message":
        return log.error_message if log.error_message else ""
    elif field == "created_at":
        return log.created_at.isoformat() if log.created_at else ""
    else:
        logger.warning(f"Field '{field}' not found in log, returning empty string")
        return ""


def get_available_fields() -> dict:
    """
    Get list of available fields for export
    
    Returns:
        Dict mapping field names to display labels
    """
    return AVAILABLE_FIELDS.copy()
